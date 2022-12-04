from openpyxl import load_workbook

def escreve_relatorio(linha_registro):
    wb = load_workbook(filename='teste.xlsx')
    ws = wb.active

    print(ws.cell(column=1, row=1).value)

    ws.cell(column=1, row=2, value=linha_registro)

    print(ws.cell(column=1, row=2).value)

    wb.save('testando.xlsx')


if __name__ == '__main__':
    escreve_relatorio('Olá')