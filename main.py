from datetime import datetime
from subprocess import call
from threading import Thread
from tkinter import Tk, Toplevel, StringVar
from tkinter.messagebox import askyesno
from os.path import abspath

from openpyxl import load_workbook, Workbook

from Modulos.construtor import Construtor


class Inicializador(Tk):
    top_level = None
    linha_registro = None
    arq = 'C:/Users/ed053781/OneDrive - intelbras.com.br/Planejamento de Treinamentos/Relatorios/Dias de fretado.xlsx'
    file = abspath(arq)
    cols = dict(data=1, ida=2, volta=3)

    def __init__(self):
        Tk.__init__(self)
        self.configura_tela()
        self.inicia_widgets()

        self.after(ms=500, func=lambda: Thread(target=self.inicia_analise_do_dia).start())

        self.mainloop()

    def configura_tela(self):
        # Parâmetros de  tamanh da tela
        tam_x, tam_y = 425, 425
        pos_x = int((self.winfo_screenwidth() - tam_x) / 2)
        pos_y = int((self.winfo_screenheight() - tam_y) / 2)
        self.geometry(f'{tam_x}x{tam_y}+{pos_x}+{pos_y}')
        self.title('Verificador offline')
        self.configure(bg='green')
        # Seta como fixo os tamanhos da tela
        self.resizable(False, False)
        self.bind('<Escape>', lambda e: self.quit())

    def inicia_widgets(self):
        scrolls = dict()
        scrolls['x_scroll'] = {
            'param': {'master': self, 'orient': 'horizontal'},
            'pack': {'side': 'bottom', 'fill': 'x'}
        }

        scrolls['y_scroll'] = {
            'param': {'master': self, 'orient': 'vertical'},
            'pack': {'side': 'right', 'fill': 'y'}
        }
        Construtor.scroll_bar(scrolls)

        relatorio = dict()
        relatorio['relatorio'] = {
            'param': {'master': self, 'bg': 'green', 'font': 'Arial 12', 'xscrollcommand': self.xscroll.set,
                      'yscrollcommand': self.yscroll.set, 'state': 'disabled'},
            'pack': {'side': 'left', 'fill': 'both'}
        }
        Construtor.text(relatorio)

    def inicia_analise_do_dia(self):
        self.escreve('Inicializando...')
        self.escreve(f'Data: {self.data_atual} - {self.dia_semana}')
        self.escreve(f'Hora: {self.hora_atual}')

        if self.dia_da_semana_de_trabalho and self.horario_de_trabalho:
            # if not self.ping:
            if not self.ping:
                self.escreve('Iniciando hora de trabalho...')
                self.abre_programas()
                self.relatorio_transporte()
            else:
                self.escreve('Verificando situação de home com usuário...')
                if askyesno('Iniciar dia de trabalho manualmente?', 'Iniciar como home & office?'):
                    self.abre_programas()
                self.encerra()
        else:
            self.escreve(info='FORA DE HORÁRIO DE TRABALHO...')
            self.encerra()

    def escreve(self, info: str):
        self.relatorio.configure(state='normal')
        self.relatorio.insert('end', info + '\n')
        self.relatorio.configure(state='disabled')

    @property
    def xscroll(self):
        return self.children['x_scroll']

    @property
    def yscroll(self):
        return self.children['x_scroll']

    @property
    def relatorio(self):
        return self.children['relatorio']

    @property
    def data_hora_atual(self):
        return datetime.now()

    @property
    def data_atual(self):
        return self.data_hora_atual.strftime('%d/%m/%Y')

    @property
    def dia_semana(self):
        dias_da_semana = {
            0: 'Segunda-feira',
            1: 'Terça-feira',
            2: 'Quarta-feira',
            3: 'Quinta-feira',
            4: 'Sexta-feira',
            5: 'Sábado',
            6: 'Domingo'
        }
        return dias_da_semana[self.data_hora_atual.weekday()]

    @property
    def hora_atual(self):
        return self.data_hora_atual.strftime('%H:%M')

    @staticmethod
    def transforma_hora(hora: str):
        return datetime.strptime(hora, '%H:%M')

    @property
    def inicio_exped(self):
        return self.transforma_hora('07:30')

    @property
    def fim_exped(self):
        return self.transforma_hora('17:30')

    @property
    def horario_de_trabalho(self):
        return self.inicio_exped < self.transforma_hora(self.hora_atual) < self.fim_exped

    @property
    def dia_da_semana_de_trabalho(self):
        return self.data_hora_atual.weekday() < 5

    @property
    def ping(self):
        hosts = ['10.100.43.1', '10.1.105.1']
        for host in hosts:
            self.escreve(f'Pingando {host}')
            ping = call(f'ping -w 300 {host}', stdout=False)
            if not ping:
                self.escreve(f'Host {host} encontrado')
                return True
            self.escreve(f'Host {host} não encontrado')
        return False

    def abre_programas(self):
        for programa in self.inicia_programas:
            call(programa, shell=True)

    @property
    def inicia_programas(self):
        with open('./Modulos/programas') as arquivo:
            programas = [link.strip('\n') for link in arquivo.readlines()]
            arquivo.close()
        return programas

    def relatorio_transporte(self):
        self.escreve_relatorio('data', self.data_atual)
        self.verifica_transporte('ida')

    def verifica_transporte(self, sentido: str):
        background = 'lightgreen'
        top_level = Toplevel(master=self)
        # Parâmetros de  tamanho da tela
        tam_x, tam_y = 450, 250
        # Parâmetros de  posição da tela
        pos_x = int((self.winfo_screenwidth() - tam_x) / 1.1)
        pos_y = int((self.winfo_screenheight() - tam_y) / 2)
        # configura tamanho e posição da tela
        top_level.geometry(f'+{pos_x}+{pos_y}')
        top_level.configure(bg=background)
        # Seta como fixo os tamanhos da tela
        top_level.resizable(False, False)
        top_level.title(f'Transporte de {sentido}')
        self.top_level = top_level

        labels = dict()
        entrys = dict()
        radio_buttons = dict()
        buttons = dict()

        linha = StringVar()

        base_param = dict(master=top_level, bg=background, font='Arial 12', )
        buttons_param = dict(activebackground=background, variable=linha, **base_param.copy(), anchor='center')

        labels['questao'] = {
            'param': dict(**base_param, text='Selecione o transporte:'),
            'grid': dict(padx=10, pady=10, columnspan=3, row=0, sticky='w')
        }

        radio_buttons['bela_vista'] = {
            'param': dict(**buttons_param, text='Bela vista - Palhoça', value='Bela vista - Palhoça',
                          command=self.desativa_entry),
            'grid': dict(padx=10, pady=10, column=0, row=1, sticky='w')
        }

        radio_buttons['sao_sebastiao'] = {
            'param': dict(**buttons_param, text='São sebastião - Palhoça', value='São sebastião - Palhoça',
                          command=self.desativa_entry),
            'grid': dict(padx=10, pady=10, column=0, row=2, sticky='w')
        }

        entrys['descricao_outros'] = {
            'param': dict(**base_param),
            'grid': dict(padx=10, pady=10, column=1, row=3, sticky='ew')
        }

        radio_buttons['outros'] = {
            'param': dict(**buttons_param, text='São sebastião - Palhoça', value='outros',
                          command=self.ativa_entry),
            'grid': dict(padx=10, pady=10, column=0, row=3, sticky='w')
        }

        buttons['registro'] = {
            'param': dict(master=top_level ,text='Registrar', command=lambda: self.registra(linha, sentido)),
            'grid': dict(padx=10, pady=10, columnspan=3, row=4, sticky='ew')
        }

        Construtor.label(labels)
        Construtor.entry(entrys)
        Construtor.radio_button(radio_buttons)
        Construtor.button(buttons)

        top_level.children['bela_vista'].invoke()

    def ativa_entry(self):
        self.top_level.children['descricao_outros'].configure(state='normal', bg='white')
        self.top_level.children['descricao_outros'].focus_force()

    def desativa_entry(self):
        self.top_level.children['descricao_outros'].delete(0, 'end')
        self.top_level.children['descricao_outros'].configure(state='readonly', bg='gray30')

    def registra(self, linha, sentido):
        info = linha.get()
        if info == 'outros':
            linha_registro = self.top_level.children['descricao_outros'].get()
            self.escreve_relatorio(sentido, linha_registro)
        else:
            linha_registro = info
            self.escreve_relatorio(sentido, linha_registro)

        if sentido == 'volta':
            self.quit()
        else:
            self.top_level.destroy()
            self.verifica_transporte('volta')

    def escreve_relatorio(self, sentido, registro):
        row = self.verifica_row(sentido)
        wb = load_workbook(filename=self.file)
        ws = wb.active

        ws.cell(column=self.cols[sentido], row=row, value=registro)
        print(ws.cell(column=self.cols[sentido], row=row).value)
        wb.save(self.file)

    def verifica_row(self, sentido):
        wb = load_workbook(filename=self.file, read_only=True)
        ws = wb.active
        for row in range(2, 30):
            val = ws.cell(column=self.cols[sentido], row=row).value
            if val == self.data_atual:
                self.escreve(info='Já existe informação de transporte para esta data')
                self.encerra()
            elif not val:
                return row

    def encerra(self):
        self.bind('<Key>', lambda e: self.quit())
        self.escreve('Aperte qualquer tecla para finalizar')


if __name__ == '__main__':
    Inicializador()

